from datetime import date, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side

__author__ = 'REBECA GONZÁLEZ BALADO'


# calcula la fecha y hora actuales
def ahora():
    return datetime.now().strftime('[%d/%m/%Y - %H:%M:%S]: ')


# genera logs con fecha y hora
def log(mensaje):
    print(ahora() + mensaje)


# genera logs con fecha y hora y deja un espacio despues
def lognw(mensaje):
    print(ahora() + mensaje + '\n')


# genera logs con fecha y hora y deja un espacio delante
def nwlog(mensaje):
    print('\n' + ahora() + mensaje)


# lee una columna hasta encontrar un valor vacio y calcula la suma de los valores del mes actual
def suma_columna_procesados(ruta, fila_inicial, columna):
    try:
        libro = load_workbook(filename=ruta)
        hoja = libro.active

        suma_total = 0
        fila = fila_inicial
        while True:

            # leemos el valor de la hoja
            suma_temp = hoja[columna + str(fila)].value

            # si esta vacio, paramos la lectura
            if suma_temp is None:
                break

            # sumamos al total y cambiamos de fila
            else:
                # si la fecha es del mes y año actuales, sumamos el valor
                if (datetime.date(hoja['A' + str(fila)].value).year == date.today().year) and \
                        (datetime.date(hoja['A' + str(fila)].value).month == date.today().month):
                    suma_total += suma_temp
                fila += 1

        return suma_total

    except TypeError:
        log('Error: Los datos a sumar no son válidos.')


# procesa datos de una lista de facturas y les da formato
def generar_facturas_procesadas(lista_facturas, titulo):
    if len(lista_facturas) == 0:
        log('No existen facturas para grabar.')

    else:
        # generamos un nuevo libro
        libro = Workbook()
        hoja = libro.active

        # generamos la cabecera
        hoja.append(['Fecha', 'Total', 'IVA'])

        # añadimos las filas con valores
        for factura in lista_facturas:
            hoja.append([factura.fecha.date(), factura.total, factura.iva])

        # ajustamos el tamaño de las celdas
        hoja.column_dimensions['A'].width = 12
        hoja.column_dimensions['B'].width = 12
        hoja.column_dimensions['C'].width = 12
        hoja.row_dimensions[1].height = 25

        # generamos el estilo
        estilo = NamedStyle(name='estilo')
        estilo.font = Font(name='Arial', size=10)
        estilo.alignment = Alignment(horizontal='left', vertical='center')
        estilo.fill = PatternFill(fill_type='solid', start_color='FEFEBD', end_color='FEFEBD')
        estilo.border = Border(left=Side(style='thin', color='999999'), right=Side(style='thin', color='999999'),
                               top=Side(style='thin', color='999999'), bottom=Side(style='thin', color='999999'))
        estilo.number_format = '#,##0.00" €"'

        # aplicamos el estilo general
        for fila in hoja['A1:C' + str(len(lista_facturas) + 1)]:
            for celda in fila:
                celda.style = estilo

        # aplicamos el estilo de la cabecera
        for fila in hoja['A1:C1']:
            for celda in fila:
                celda.alignment = Alignment(horizontal='center', vertical='center')
                celda.fill = PatternFill(fill_type='solid', start_color='D9D982', end_color='D9D982')

        # aplicamos el formato de fecha
        for fila in hoja['A2:A' + str(len(lista_facturas) + 1)]:
            for celda in fila:
                celda.number_format = 'DD/MM/YYYY'

        # escribimos los titulos apropiados
        hoja.title = titulo + '-procesadas'
        libro.save('../datos/facturas/' + titulo + '-procesadas.xlsx')
        log('Se ha generado el fichero de ' + titulo + ' procesadas.')


# calcula el coste para la empresa a partir de una nomina
def calcular_gasto_nomina(fichero):
    try:
        libro = load_workbook(filename='../datos/nominas/' + fichero)
        hoja = libro.active

        log('Se ha abierto el fichero ' + fichero + '.')

        # leemos de la hoja los valores primitivos
        d12 = hoja['D12'].value  # numero pagos salario
        e12 = hoja['E12'].value  # salario mensual
        f16 = hoja['F16'].value  # incentivos
        d17 = hoja['D17'].value  # numero pagas extraordinarias
        e42 = hoja['E42'].value  # porcentaje contingencias comunes
        e43 = hoja['E43'].value  # porcentaje contingencias profesionales
        e44 = hoja['E44'].value  # porcentaje cotizacion horas extraordinarias

        # realizamos la primera capa de formulas
        f12 = d12 * e12  # salario total
        f17 = d17 * e12  # pagas extraordinarias totales

        # realizamos la segunda capa de formulas
        f23 = f12 + f16 + f17  # total devengado

        # realizamos la tercera capa de formulas
        f42 = f23 * e42  # total contingencias comunes
        f43 = f23 * e43  # total contingencias profesionales
        f44 = f23 * e44  # total horas extraordinarias

        # realizamos el calculo final
        f46 = f42 + f43 + f44 + f23  # coste total empresa
        coste_empresa = f46

        # devolvemos el gasto
        return coste_empresa

    except TypeError:
        log('Error: Los datos de la nómina no son válidos.')


# graba los datos introducidos segun el mes, un rango de columnas y unos valores
def grabar_informe_mensual(ruta, rango_columnas, filas_valores):
    libro = load_workbook(filename=ruta)
    hoja = libro.active

    # agregamos el rango de columnas de enero a diciembre en una lista
    letras_columnas = rango_columnas.split()
    # conseguimos la letra de columna del mes actual
    letra_mes = letras_columnas[date.today().month - 1]

    # grabamos los datos en la columna correspondiente al mes actual
    for key in filas_valores.keys():
        hoja[letra_mes + str(key)] = filas_valores[key]

    # sobreescribimos el fichero
    libro.save(filename=ruta)

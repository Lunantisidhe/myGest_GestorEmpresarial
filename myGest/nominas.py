import calendar
from datetime import date
from openpyxl import load_workbook
from recursos import gestor
from recursos.clases import Empleado

__author__ = 'REBECA GONZÁLEZ BALADO'

lista_empleados = []
tabla_salarial = {}


# carga los valores de la tabla de salarios
def cargar_tabla_salarial():
    try:
        libro = load_workbook(filename='../datos/rrhh/empleados.xlsx')
        hoja = libro['tablasSalariales']

        fila = 2
        while True:
            categoria = hoja['A' + str(fila)].value
            salario = hoja['B' + str(fila)].value

            # si algun parametro esta vacio, detenemos la busqueda
            if (categoria is None) or (salario is None):
                break

            # si los valores de la fila son correctos, la grabamos en la tabla salarial y cambiamos de fila
            if (type(categoria) is str) and ((type(salario) is int) or (type(salario) is float)):
                tabla_salarial[categoria] = salario
            fila += 1

        if len(tabla_salarial) == 0:
            gestor.log('Error: Error al cargar los datos de la tabla salarial.')
        else:
            gestor.log('Se ha cargado la tabla salarial.')

    except FileNotFoundError:
        gestor.log('Error: No se ha encontrado el fichero de empleados.')
    except KeyError:
        gestor.log('Error: No se ha encontrado la hoja de la tabla salarial.')


# calcula el periodo de liquidacion (primer y ultimo dia del mes)
def calcular_periodo_liquidacion():
    dia_actual = date.today()
    primer_dia_mes = dia_actual.replace(day=1)
    ultimo_dia_mes = dia_actual.replace(day=calendar.monthrange(dia_actual.year, dia_actual.month)[1])

    return primer_dia_mes, ultimo_dia_mes


# lee los datos de los empleados del fichero de empleados
def leer_datos_empleado():
    try:
        libro = load_workbook(filename='../datos/rrhh/empleados.xlsx')
        hoja = libro['empleados']

        fila = 2
        while True:
            # leemos los valores de la hoja
            nombre = hoja['A' + str(fila)].value
            categoria = hoja['B' + str(fila)].value
            numero_cuenta = hoja['C' + str(fila)].value
            incentivos = hoja['D' + str(fila)].value

            # si alguno esta vacio (salvo incentivos), paramos la lectura
            if (nombre is None) or (categoria is None) or (numero_cuenta is None):
                break

            # guardamos los datos en la lista de empleados y cambiamos de fila
            if (type(nombre) is str) and (type(categoria) is str) and (type(numero_cuenta) is str) and \
                    ((type(incentivos) is int) or (type(incentivos) is float) or (incentivos is None)) and \
                    (tabla_salarial.get(categoria) is not None):
                lista_empleados.append(Empleado(nombre, categoria, calcular_periodo_liquidacion(),
                                                tabla_salarial.get(categoria), incentivos, 2, numero_cuenta))
            fila += 1

        gestor.lognw('Se han leído los datos de ' + str(len(lista_empleados)) + ' empleados.')

    except FileNotFoundError:
        gestor.lognw('Error: No se ha encontrado el fichero de empleados.')
    except KeyError:
        gestor.lognw('Error: No se ha encontrado la hoja de empleados.')
    except ValueError:
        gestor.lognw('Error: Los datos del fichero de empleados no son válidos.')


# genera una nomina por cada empleado
def generar_nominas():
    if len(lista_empleados) == 0:
        gestor.log('No existen empleados para grabar.')

    elif len(tabla_salarial) == 0:
        gestor.log('Error: No se pueden generar las nóminas al no existir la tabla salarial.')

    else:
        try:
            for empleado in lista_empleados:
                # importamos la plantilla
                libro = load_workbook(filename='../datos/nominas/plantilla.xlsx')

                # abrimos la hoja
                hoja = libro.active

                # grabamos los datos del empleado
                hoja['E2'] = empleado.nombre
                hoja['F5'] = empleado.categoria
                hoja['C8'] = empleado.periodo_liquidacion[0]
                hoja['D8'] = empleado.periodo_liquidacion[1]
                hoja['E12'] = empleado.salario
                if empleado.incentivos is not None:
                    hoja['F16'] = empleado.incentivos
                hoja['D17'] = empleado.pagas_extra
                hoja['C6'] = empleado.numero_cuenta

                # guardamos los cambios
                libro.save(filename='../datos/nominas/nomina_' + empleado.nombre.replace(' ', '') + '.xlsx')
                gestor.log('La nómina de ' + empleado.nombre + ' ha sido generada satisfactoriamente.')

            gestor.nwlog('Se han generado todas las nóminas de empleado.')

        except FileNotFoundError:
            gestor.nwlog('Error: No se ha encontrado la plantilla de nóminas.')


def main():
    cargar_tabla_salarial()
    leer_datos_empleado()
    generar_nominas()


if __name__ == '__main__':
    main()
